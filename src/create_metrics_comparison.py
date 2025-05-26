import pandas as pd
import os
import glob
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Definir las métricas que queremos analizar
METRICS = {
    'F1-Score (Micro)': 'F1-Score (Micro)',
    'F1-Score (Macro)': 'F1-Score (Macro)',
    'F1-Score (Samples)': 'F1-Score (Samples)',
    'Accuracy': 'Accuracy (Subset)',
    'Hamming Loss': 'Hamming Loss'
}

# Definir el orden específico de los datasets
DATASET_ORDER = [
    'Emotions',
    'CHD_49',
    'VirusGo',
    'Foodtruck',
    'Water-quality',
    'Birds',
    'Yahoo_Arts',
    'Yahoo_Business',
    'Mediamill',
    'Bibtex'
]

# Definir los modelos y sus carpetas
MODELS = {
    'CC': 'CC',
    'ECC': 'ECC',
    'FCCC': 'FCCC',
    'LCC-MLC': 'LCC-MLC'
}

# Función para leer los datos de un modelo
def read_model_data(model_folder):
    model_path = os.path.join('experimentos/Moyano', model_folder)
    metrics_data = {}
    
    # Encontrar todos los archivos CSV de métricas (excluyendo global_metrics.csv)
    csv_files = glob.glob(os.path.join(model_path, '*_metrics.csv'))
    csv_files = [f for f in csv_files if 'global_metrics.csv' not in f]
    
    for csv_file in csv_files:
        # Extraer el nombre del dataset del nombre del archivo
        dataset = os.path.basename(csv_file).replace('_metrics.csv', '')
        
        # Leer el CSV
        df = pd.read_csv(csv_file)
        
        # Obtener el promedio de cada métrica
        metrics_data[dataset] = {}
        for metric_name, csv_column in METRICS.items():
            metrics_data[dataset][metric_name] = df[csv_column].iloc[-1]  # Última fila es el promedio
    
    return metrics_data

def highlight_best_values(worksheet, df, metric_name, start_row=2):
    """Marca en negrita los mejores valores de cada fila"""
    find_min = metric_name == 'Hamming Loss'

    for row in range(start_row, start_row + len(df)):
        values = []
        for col in range(2, 2 + len(MODELS)):  # Columnas de datos
            cell = worksheet.cell(row=row, column=col)
            val = cell.value
            try:
                num = float(val)
                values.append((num, col))
            except (ValueError, TypeError):
                print(f"Valor inválido en fila {row}, columna {col}: {val}")

        if values:
            best_value = min(values, key=lambda x: x[0]) if find_min else max(values, key=lambda x: x[0])
            best_cell = worksheet.cell(row=row, column=best_value[1])
            best_cell.font = Font(bold=True)

def add_mean_row(worksheet, df, metric_name, start_row):
    """Añade una fila con las medias y marca la mejor en negrita"""
    # Calcular medias
    means = df.mean()
    
    # Escribir "Mean" en la primera columna
    worksheet.cell(row=start_row, column=1, value="Mean")
    
    # Escribir las medias y encontrar la mejor
    find_min = metric_name == 'Hamming Loss'
    best_col = None
    best_value = float('inf') if find_min else float('-inf')
    
    for col, (model, mean) in enumerate(means.items(), start=2):
        cell = worksheet.cell(row=start_row, column=col, value=mean)
        if find_min:
            if mean < best_value:
                best_value = mean
                best_col = col
        else:
            if mean > best_value:
                best_value = mean
                best_col = col
    
    # Marcar la mejor media en negrita
    if best_col:
        worksheet.cell(row=start_row, column=best_col).font = Font(bold=True)

def main():
    # Crear un diccionario para almacenar todos los datos
    all_data = {}
    
    # Leer datos de cada modelo
    for model_name, model_folder in MODELS.items():
        all_data[model_name] = read_model_data(model_folder)
    
    # Crear un Excel con una hoja por cada métrica
    with pd.ExcelWriter('resultados/Metric_x_Dataset.xlsx', engine='openpyxl') as writer:
        for metric_name in METRICS.keys():
            # Crear un DataFrame para esta métrica
            metric_data = []
            
            # Usar el orden específico de datasets
            for dataset in DATASET_ORDER:
                row = {'Dataset': dataset}
                for model_name in MODELS.keys():
                    if dataset in all_data[model_name]:
                        row[model_name] = all_data[model_name][dataset][metric_name]
                    else:
                        row[model_name] = None
                metric_data.append(row)
            
            # Convertir a DataFrame
            df = pd.DataFrame(metric_data)
            df.set_index('Dataset', inplace=True)
            
            # Guardar en una hoja del Excel
            df.to_excel(writer, sheet_name=metric_name)
            
            # Obtener la hoja actual
            worksheet = writer.sheets[metric_name]
            
            # Marcar los mejores valores en negrita
            highlight_best_values(worksheet, df, metric_name)
            
            # Añadir fila de medias
            add_mean_row(worksheet, df, metric_name, len(df) + 2)

if __name__ == '__main__':
    main() 